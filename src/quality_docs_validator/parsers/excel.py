"""Excel (.xlsx) parser for PFMEA and Control Plan documents.

MVP behaviour (see docs/DECISIONS.md D1/D2): `.xlsx` only, recommended template plus a small set
of column aliases. The first worksheet is read. The header row does not have to be the very first
row: the first row (within the first few) that contains an `operation_id`-aliased column is used as
the header, so a leading title/blank row is tolerated. Headers are normalised (lowercased,
non-alphanumeric stripped) before being matched against the alias maps. Unknown/extra columns are
ignored. Configurable multi-template mapping is intentionally out of MVP scope (v0.2).
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from ..models import ControlPlanRow, PFMEARow


class ParseError(Exception):
    """Raised when a file cannot be parsed into the expected document shape."""


# Canonical field -> accepted normalised header aliases. Headers are normalised (lowercased,
# non-alphanumeric stripped) before matching, so "Operation No." == "operationno" and "Op #" == "op".
# Aliases are deliberately specific; ambiguous generic tokens (control, method, description, number,
# id, status) are not used as new aliases to avoid mis-mapping columns.
PFMEA_ALIASES: dict[str, set[str]] = {
    "operation_id": {
        "operationid", "opid", "opno", "operationno", "operation", "processnumber",
        "operationnumber", "opnumber", "op", "processno", "stepno",
    },
    "process_step": {
        "processstep", "processfunction", "operationdescription", "step", "processname",
        "processdescription", "stepdescription",
    },
    "failure_mode": {"failuremode", "potentialfailuremode", "failure"},
    "effect": {"effect", "potentialeffect", "effectsoffailure"},
    "severity": {"severity", "sev", "s", "severityrating"},
    "cause": {"cause", "potentialcause", "causeoffailure"},
    "prevention_control": {"preventioncontrol", "currentpreventioncontrol", "prevention"},
    "detection_control": {
        "detectioncontrol", "currentdetectioncontrol", "currentcontrolsdetection", "detectionmethod",
    },
    "detection": {"detection", "det", "d"},
    "special_characteristic": {
        "specialcharacteristic", "specialcharacteristics", "specialchar", "classification", "sc",
        "keycharacteristic", "criticalcharacteristic",
    },
}

CONTROL_PLAN_ALIASES: dict[str, set[str]] = {
    "operation_id": {
        "operationid", "opid", "opno", "operationno", "operation", "processnumber",
        "operationnumber", "opnumber", "op", "processno", "stepno",
    },
    "process_step": {
        "processstep", "processname", "processdescription", "operationdescription", "step",
        "stepdescription",
    },
    "characteristic": {"characteristic", "productcharacteristic", "processcharacteristic"},
    "special_characteristic": {
        "specialcharacteristic", "specialcharacteristics", "specialchar", "classification", "sc",
        "keycharacteristic", "criticalcharacteristic",
    },
    "control_method": {
        "controlmethod", "control", "method", "evaluationmeasurementtechnique",
        "controltechnique", "measurementtechnique",
    },
    "detection_method": {"detectionmethod", "inspectionmethod", "evaluationmethod"},
    "reaction_plan": {
        "reactionplan", "reaction", "responseplan", "correctiveaction", "outofcontrolaction",
    },
}

_TRUE_TOKENS = {"yes", "y", "true", "1", "x", "cc", "sc", "critical", "significant", "*", "✓", "✔"}


def _normalise(header: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(header).strip().lower()) if header is not None else ""


def _build_header_map(headers: list[object], aliases: dict[str, set[str]]) -> dict[str, int]:
    """Map canonical field name -> column index, using the alias table."""
    field_to_index: dict[str, int] = {}
    for index, header in enumerate(headers):
        norm = _normalise(header)
        if not norm:
            continue
        for field, accepted in aliases.items():
            if norm in accepted and field not in field_to_index:
                field_to_index[field] = index
    return field_to_index


def _cell(row: tuple, index: int | None) -> object:
    if index is None or index >= len(row):
        return None
    return row[index]


def _text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_int(value: object) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return None


def _to_bool(value: object) -> bool:
    if value is None:
        return False
    return _normalise(value) in {_normalise(t) for t in _TRUE_TOKENS}


HEADER_SCAN_ROWS = 10  # how many leading rows to scan when locating the header row


def _load_rows(path: Path) -> list[tuple]:
    if not path.exists():
        raise ParseError(f"File not found: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ParseError(
            f"Only .xlsx files are supported in the MVP (got '{path.suffix or 'no extension'}'). "
            f"CSV and other formats are planned for v0.2 (see docs/ROADMAP.md)."
        )
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a variety of types for corrupt/non-xlsx files
        raise ParseError(f"Could not open '{path.name}' as an .xlsx workbook: {exc}") from exc
    sheet = workbook.active
    if sheet is None:
        workbook.close()
        raise ParseError(f"Workbook '{path.name}' has no worksheets.")
    rows = [r for r in sheet.iter_rows(values_only=True) if any(c is not None for c in r)]
    workbook.close()
    if not rows:
        raise ParseError(f"Worksheet in '{path.name}' is empty (no rows with content).")
    return rows


def _find_header(
    rows: list[tuple], aliases: dict[str, set[str]], doc: str, path: Path
) -> tuple[int, dict[str, int]]:
    """Locate the header row: the first scanned row that exposes an `operation_id` column."""
    for index, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        field_map = _build_header_map(list(row), aliases)
        if "operation_id" in field_map:
            return index, field_map
    seen = ", ".join(str(c) for c in rows[0] if c is not None) or "(blank)"
    raise ParseError(
        f"Could not find an 'operation_id' column in {doc} '{path.name}'. "
        f"Looked at the first {min(HEADER_SCAN_ROWS, len(rows))} row(s); first row was: {seen}. "
        f"Use the recommended template / a recognised column alias (docs/DECISIONS.md D2, "
        f"docs/FINDINGS.md)."
    )


def parse_pfmea(path: str | Path) -> list[PFMEARow]:
    path = Path(path)
    rows = _load_rows(path)
    header_index, field_map = _find_header(rows, PFMEA_ALIASES, "PFMEA", path)
    data = rows[header_index + 1 :]

    out: list[PFMEARow] = []
    for offset, raw in enumerate(data, start=header_index + 2):
        operation_id = _text(_cell(raw, field_map.get("operation_id")))
        if operation_id is None:
            continue  # skip blank/spacer rows
        out.append(
            PFMEARow(
                operation_id=operation_id,
                process_step=_text(_cell(raw, field_map.get("process_step"))),
                failure_mode=_text(_cell(raw, field_map.get("failure_mode"))),
                effect=_text(_cell(raw, field_map.get("effect"))),
                severity=_to_int(_cell(raw, field_map.get("severity"))),
                cause=_text(_cell(raw, field_map.get("cause"))),
                prevention_control=_text(_cell(raw, field_map.get("prevention_control"))),
                detection_control=_text(_cell(raw, field_map.get("detection_control"))),
                detection=_to_int(_cell(raw, field_map.get("detection"))),
                special_characteristic=_to_bool(_cell(raw, field_map.get("special_characteristic"))),
                row_number=offset,
            )
        )
    if not out:
        raise ParseError(f"No PFMEA data rows found in '{path.name}'.")
    return out


def parse_control_plan(path: str | Path) -> list[ControlPlanRow]:
    path = Path(path)
    rows = _load_rows(path)
    header_index, field_map = _find_header(rows, CONTROL_PLAN_ALIASES, "Control Plan", path)
    data = rows[header_index + 1 :]

    out: list[ControlPlanRow] = []
    for offset, raw in enumerate(data, start=header_index + 2):
        operation_id = _text(_cell(raw, field_map.get("operation_id")))
        if operation_id is None:
            continue
        out.append(
            ControlPlanRow(
                operation_id=operation_id,
                process_step=_text(_cell(raw, field_map.get("process_step"))),
                characteristic=_text(_cell(raw, field_map.get("characteristic"))),
                special_characteristic=_to_bool(_cell(raw, field_map.get("special_characteristic"))),
                control_method=_text(_cell(raw, field_map.get("control_method"))),
                detection_method=_text(_cell(raw, field_map.get("detection_method"))),
                reaction_plan=_text(_cell(raw, field_map.get("reaction_plan"))),
                row_number=offset,
            )
        )
    if not out:
        raise ParseError(f"No Control Plan data rows found in '{path.name}'.")
    return out
