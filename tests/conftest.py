"""Shared test fixtures: builds small synthetic .xlsx workbooks on the fly."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

EXAMPLES_DIR = Path(__file__).resolve().parents[1] / "examples"


def _make_xlsx(path: Path, headers: list[str], rows: list[list]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def _make_multi_sheet_xlsx(path: Path, sheets: dict[str, tuple[list, list]]) -> Path:
    """Build a workbook with several named sheets: {name: (headers, rows)}."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    for name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(title=name)
        ws.append(headers)
        for row in rows:
            ws.append(row)
    wb.save(path)
    return path


@pytest.fixture
def make_xlsx():
    return _make_xlsx


@pytest.fixture
def make_multi_sheet_xlsx():
    return _make_multi_sheet_xlsx


@pytest.fixture(scope="session")
def example_files() -> tuple[Path, Path]:
    """The committed synthetic examples; regenerate them if missing."""
    pfmea = EXAMPLES_DIR / "pfmea.xlsx"
    control_plan = EXAMPLES_DIR / "control-plan.xlsx"
    if not (pfmea.exists() and control_plan.exists()):
        from examples.generate_examples import main as generate  # type: ignore

        generate()
    return pfmea, control_plan
